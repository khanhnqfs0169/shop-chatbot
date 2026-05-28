"""
Chay file nay MOT LAN de tao file products.xlsx mau.
Sau do ban mo file do va chinh sua san pham that cua shop.

Cai dat: pip install openpyxl
Chay:    python products_sample.py
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

SAMPLE_PRODUCTS = [
    {"Ma SP": "AT001", "Ten san pham": "Ao thun basic unisex", "Loai": "Ao thun",
     "Size co san": "S, M, L, XL, XXL", "Mau sac": "Trang, Den, Xam, Navy, Be",
     "Gia (VND)": 150000, "Gia sale": 120000, "Chat lieu": "Cotton 100%",
     "Mo ta": "Form rong, thoang mat, phu hop mac hang ngay", "Con hang": "Co"},
    {"Ma SP": "AT002", "Ten san pham": "Ao thun graphic in hoa tiet", "Loai": "Ao thun",
     "Size co san": "S, M, L, XL", "Mau sac": "Trang, Den",
     "Gia (VND)": 180000, "Gia sale": "", "Chat lieu": "Cotton 65% Polyester 35%",
     "Mo ta": "In hoa tiet noi bat, unisex, form regular fit", "Con hang": "Co"},
    {"Ma SP": "PL001", "Ten san pham": "Ao polo nam basic", "Loai": "Ao polo",
     "Size co san": "M, L, XL, XXL", "Mau sac": "Trang, Den, Xanh navy, Do do",
     "Gia (VND)": 250000, "Gia sale": 200000, "Chat lieu": "Ca sau cotton",
     "Mo ta": "Form slim fit, co be, phu hop di lam hoac dao pho", "Con hang": "Co"},
    {"Ma SP": "CN001", "Ten san pham": "Ao khoac chong nang nu", "Loai": "Ao khoac",
     "Size co san": "S, M, L", "Mau sac": "Den, Trang, Hong nhat, Xanh mint",
     "Gia (VND)": 320000, "Gia sale": 280000, "Chat lieu": "Vai du chong tia UV",
     "Mo ta": "Chong tia UV 99%, nhe, co tui keo khoa, phu hop di xe may", "Con hang": "Co"},
    {"Ma SP": "QJ001", "Ten san pham": "Quan jeans nam slim fit", "Loai": "Quan jeans",
     "Size co san": "29, 30, 31, 32, 33, 34", "Mau sac": "Xanh nhat, Xanh dam, Den",
     "Gia (VND)": 450000, "Gia sale": 380000, "Chat lieu": "Denim co gian 4 chieu",
     "Mo ta": "Form om vua, ton dang, phu hop nhieu hoan canh", "Con hang": "Co"},
    {"Ma SP": "QST001", "Ten san pham": "Quan short the thao", "Loai": "Quan short",
     "Size co san": "S, M, L, XL", "Mau sac": "Den, Xam, Xanh lam",
     "Gia (VND)": 200000, "Gia sale": "", "Chat lieu": "Polyester thoang khi",
     "Mo ta": "Phu hop tap gym, chay bo, mac nha. Co tui keo khoa", "Con hang": "Co"},
]

def create_products_excel():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "San pham"
    headers = list(SAMPLE_PRODUCTS[0].keys())
    header_fill = PatternFill(start_color="2196F3", end_color="2196F3", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row_idx, product in enumerate(SAMPLE_PRODUCTS, 2):
        for col_idx, key in enumerate(headers, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=product[key])
            cell.alignment = Alignment(vertical="center", wrap_text=True)
    col_widths = [10, 30, 15, 25, 30, 15, 12, 25, 45, 12]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 25
    wb.save("products.xlsx")
    print("Da tao 'products.xlsx'! Mo file va nhap san pham that cua shop vao.")

if __name__ == "__main__":
    create_products_excel()
